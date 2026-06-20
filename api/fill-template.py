from http.server import BaseHTTPRequestHandler
import json
import base64
import io


def round_price(price):
    base = int(price // 100) * 100
    candidates = [base - 51, base - 1, base + 49, base + 99, base + 149]
    return min(candidates, key=lambda c: abs(c - price))


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            import openpyxl
        except ImportError:
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps({'error': 'openpyxl not available'}).encode())
            return

        content_length = int(self.headers['Content-Length'])
        body = json.loads(self.rfile.read(content_length))

        file_b64 = body['file']
        products = body['products']  # {asin: {min, max}}

        file_bytes = base64.b64decode(file_b64)
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), keep_vba=True)
        ws = wb['Template']

        # Find attribute row (contains contribution_sku)
        attr_row = None
        for row in ws.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if cell.value and 'contribution_sku' in str(cell.value):
                    attr_row = cell.row
                    break
            if attr_row:
                break
        if not attr_row:
            attr_row = 5

        # Map attribute names to column indices
        col_map = {}
        for cell in ws[attr_row]:
            if cell.value:
                col_map[str(cell.value).strip()] = cell.column

        asin_col = col_map.get('merchant_suggested_asin#1.value')
        search_term_col = col_map.get('::your_search_term')
        action_col = col_map.get('::record_action')
        recommended_col = col_map.get('::recommended_action')
        # Also check by label row (row 4)
        if not recommended_col:
            for cell in ws[4]:
                if cell.value and 'recommended' in str(cell.value).lower():
                    recommended_col = cell.column
                    break

        def set_col(row_num, col_name, value):
            idx = col_map.get(col_name)
            if idx:
                ws.cell(row=row_num, column=idx).value = value

        filled = 0
        not_in_catalog = []
        filled_asins = []
        data_start = attr_row + 2
        for row_num in range(data_start, ws.max_row + 1):
            asin_val = ws.cell(row=row_num, column=asin_col).value if asin_col else None
            # Fallback to search term column (for "Not in Amazon" rows where merchant_asin is empty)
            if not asin_val and search_term_col:
                asin_val = ws.cell(row=row_num, column=search_term_col).value
            action_val = ws.cell(row=row_num, column=action_col).value if action_col else None
            recommended_val = ws.cell(row=row_num, column=recommended_col).value if recommended_col else None

            if not asin_val:
                continue
            asin = str(asin_val).strip()

            # Detect "Not in Amazon"
            if recommended_val and 'not in amazon' in str(recommended_val).lower():
                not_in_catalog.append(asin)
                continue

            if action_val and str(action_val).strip() != 'Add Product':
                continue
            if asin not in products:
                continue

            p = products[asin]
            min_price = p['min']
            max_price = p['max']

            set_col(row_num, 'condition_type#1.value', 'New')
            set_col(row_num, 'fulfillment_availability#1.fulfillment_channel_code', 'Fulfilment by Merchant (Default)')
            set_col(row_num, 'fulfillment_availability#1.quantity', 1)
            set_col(row_num, 'fulfillment_availability#1.lead_time_to_ship_max_days', 5)
            set_col(row_num, 'purchasable_offer[marketplace_id=ARBP9OOSHTCHU][audience=ALL]#1.our_price#1.schedule#1.value_with_tax', max_price)
            set_col(row_num, 'purchasable_offer[marketplace_id=ARBP9OOSHTCHU][audience=ALL]#1.automated_pricing_merchandising_rule_plan#1.merchandising_rule.rule_id', 'Low 1 pound')
            set_col(row_num, 'purchasable_offer[marketplace_id=ARBP9OOSHTCHU][audience=ALL]#1.minimum_seller_allowed_price#1.schedule#1.value_with_tax', min_price)
            set_col(row_num, 'purchasable_offer[marketplace_id=ARBP9OOSHTCHU][audience=ALL]#1.maximum_seller_allowed_price#1.schedule#1.value_with_tax', max_price)
            set_col(row_num, 'supplier_declared_dg_hz_regulation#1.value', 'Not Applicable')
            set_col(row_num, 'supplier_declared_dg_hz_regulation#2.value', 'Not Applicable')
            set_col(row_num, 'supplier_declared_dg_hz_regulation#3.value', 'Not Applicable')
            set_col(row_num, 'supplier_declared_dg_hz_regulation#4.value', 'Not Applicable')
            set_col(row_num, 'supplier_declared_dg_hz_regulation#5.value', 'Not Applicable')
            filled_asins.append(asin)
            filled += 1

        out = io.BytesIO()
        wb.save(out)
        out_b64 = base64.b64encode(out.getvalue()).decode()

        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(json.dumps({
            'file': out_b64,
            'not_in_catalog': not_in_catalog,
            'filled_asins': filled_asins,
            'filled': filled
        }).encode())

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
